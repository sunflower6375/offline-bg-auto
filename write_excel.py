import pandas as pd
import gspread
import datetime
from dateutil.relativedelta import relativedelta
from openpyxl import load_workbook

month = datetime.datetime.now() + relativedelta(months=1)
formatted_date = month.strftime("%b'%y")
formatted_month = month.strftime("%B")

# source gsheet file
url = "https://docs.google.com/spreadsheets/d/130_NbFDUy7Mzk7sEjcyqM7nBuoEXQxCejz7zqJ77c6o/edit#gid=509276703"

# destination excel file
sheet_nm = 'Shopee Marketing Projects'
start_row = 26 
start_col = 3 
new_path = rf"C:\Users\linh.mynguyen\OneDrive - Seagroup\Linh's folder\0. Offline MKT\{formatted_date} Offline Budget\Shopee {formatted_date} Budget Template (Marketing Projects).xlsx"

def read_gsheet(formatted_date, url):
    gc = gspread.service_account(filename="credentials.json") # log in with credentials.json file
    workbook = gc.open_by_url(url) # open the gsheet
    sheet = workbook.worksheet(f"{formatted_date} - Offline Budget") # select the sheet
    raw_data = sheet.get_all_values() # obtain all values from the sheet
    data = pd.DataFrame(raw_data) # convert the data into a dataframe
    data = data.iloc[53:,2:18] # select the rows and columns needed
    data.columns = data.iloc[0] # specify the header column
    return data

def write_excel(sheet, start_row, start_col, file_path):
    wb = load_workbook(file_path) # use openpyxl to open workbook at specified path
    ws = wb[sheet] # choose the worksheet
    for col in ws.column_dimensions.values(): #unhide all columns
        col.hidden = False

    gsheet = read_gsheet(formatted_date, url) #call the function read_gsheet
    data = gsheet.values.tolist() #convert DataFrame to List type
    
    for row_index, row_data in enumerate(data): # loop though every row and its data
        for column_index, value in enumerate(row_data): # loop though every column in each row, also get the 'value'
            cell = ws.cell(row=start_row + row_index, column=start_col + column_index) # find the coresponding cell location, given start_row and start_col
            if isinstance(value, str) and value.isdigit():  # check if value is a digit stored as string
                value = int(value)  # convert to integer
            elif isinstance(value, str):
                try:
                    value = float(value)  # attempt to convert to float
                except ValueError:
                    pass
            cell.value = value # assign the cell value with 'value' above
    wb.save(new_path) # save and accept change

write_excel(sheet_nm, start_row, start_col, new_path) # execute code

